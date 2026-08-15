"""Genera los fixtures xlsx anonimizados de tests/fixtures/ (sección 24 de
docs/architecture.md: fixtures anonimizados, nunca datos reales en Git).

Uso: python scripts/generate_fixtures.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).resolve().parents[1] / "tests" / "fixtures"

HEADERS = [
    "descripcion",
    "cantidad",
    "unidad",
    "fabricante",
    "part_number",
    "norma",
    "especificacion",
    "observaciones",
]


def _write_sheet(ws, rows: list[list]) -> None:
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)


def build_rfq_simple_1_linea() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ"
    _write_sheet(
        ws,
        [["Válvula de bola API 6D 4\"", 12, "pza", None, None, "API 6D", "4 pulgadas, clase 300#", None]],
    )
    wb.save(FIXTURES_DIR / "rfq_simple_1_linea.xlsx")


def build_rfq_masiva_500_lineas() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ"
    rows = [
        [f"Producto genérico {i:04d}", (i % 20) + 1, "pza", None, None, None, None, None]
        for i in range(1, 501)
    ]
    _write_sheet(ws, rows)
    wb.save(FIXTURES_DIR / "rfq_masiva_500_lineas.xlsx")


def build_rfq_lineas_duplicadas() -> None:
    """100 tornillos M10 idénticos repartidos en 3 filas -> debe fusionarse
    en 1 sola línea con quantity=100 (caso 3 de docs/test_strategy.md)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ"
    rows = [
        ["Tornillos M10 acero inoxidable", 40, "pza", None, None, None, "acero inoxidable", None],
        ["Tornillos M10 acero inoxidable", 35, "pza", None, None, None, "acero inoxidable", None],
        ["Tornillos M10 acero inoxidable", 25, "pza", None, None, None, "acero inoxidable", None],
    ]
    _write_sheet(ws, rows)
    wb.save(FIXTURES_DIR / "rfq_lineas_duplicadas.xlsx")


def build_rfq_multitab() -> None:
    """3 hojas, solo una con datos de producto (caso 11)."""
    wb = Workbook()
    portada = wb.active
    portada.title = "Portada"
    portada.append(["Cotización solicitada por Cliente XYZ"])
    portada.append(["Fecha:", "2026-08-01"])

    productos = wb.create_sheet("Productos")
    _write_sheet(
        productos,
        [
            ["Brida 6\" 150#", 3, "pza", None, None, "ASTM A105", "6 pulgadas, clase 150#", "sin marca"],
            ["Codo 90° 4\"", 6, "pza", None, None, "ASTM A234", None, None],
        ],
    )

    notas = wb.create_sheet("Notas")
    notas.append(["Notas internas, no es parte del pedido"])

    wb.save(FIXTURES_DIR / "rfq_multitab.xlsx")


if __name__ == "__main__":
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    build_rfq_simple_1_linea()
    build_rfq_masiva_500_lineas()
    build_rfq_lineas_duplicadas()
    build_rfq_multitab()
    print(f"Fixtures generados en {FIXTURES_DIR}")
